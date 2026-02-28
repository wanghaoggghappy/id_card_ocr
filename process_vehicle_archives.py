"""
车辆文档批处理脚本
处理压缩包中的车辆文档，识别信息，输出Excel，整理文件
"""

import os
import sys
import shutil
import tempfile
import argparse
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("错误: 缺少 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)

from vehicle_document_processor import VehicleDocumentProcessor, DocumentFile
from vehicle_info_extractor import VehicleInfo, VehicleInfoExtractor


class VehicleArchiveProcessor:
    """车辆档案批处理器"""
    
    def __init__(self, output_dir: str = "output", ocr_engine: str = "paddleocr"):
        """
        初始化处理器
        
        Args:
            output_dir: 输出目录
            ocr_engine: OCR引擎
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        self.processor = VehicleDocumentProcessor(ocr_engine=ocr_engine)
        self.extractor = VehicleInfoExtractor()
    
    def _detect_subfolders(self, extract_dir: str) -> List[Path]:
        """
        检测解压目录中的子文件夹
        
        Returns:
            子文件夹路径列表，如果没有子文件夹则返回空列表
        """
        extract_path = Path(extract_dir)
        subfolders = [d for d in extract_path.iterdir() if d.is_dir()]
        
        # 过滤掉系统文件夹
        subfolders = [d for d in subfolders if not d.name.startswith('.')]
        
        return subfolders
    
    def _skip_intermediate_folder(self, extract_dir: str) -> str:
        """
        跳过中间文件夹层级
        
        如果解压后只有一个文件夹，且该文件夹下没有图片而是有多个子文件夹，
        则自动进入该文件夹作为新的根目录
        
        Args:
            extract_dir: 解压目录
            
        Returns:
            实际的工作目录
        """
        subfolders = self._detect_subfolders(extract_dir)
        has_root_images = self._has_images_in_root(extract_dir)
        
        # 如果只有一个子文件夹，且根目录没有图片
        if len(subfolders) == 1 and not has_root_images:
            potential_root = subfolders[0]
            
            # 检查这个文件夹下是否有图片
            has_images_inside = self._has_images_in_root(str(potential_root))
            
            # 检查这个文件夹下是否还有子文件夹
            sub_subfolders = self._detect_subfolders(str(potential_root))
            
            # 如果该文件夹下没有图片，但有多个子文件夹，说明这是中间层
            if not has_images_inside and len(sub_subfolders) > 0:
                print(f"  检测到中间文件夹: {potential_root.name}")
                print(f"  自动进入该文件夹，找到 {len(sub_subfolders)} 个车辆档案文件夹")
                return str(potential_root)
        
        return extract_dir
    
    def _has_images_in_root(self, extract_dir: str) -> bool:
        """
        检查根目录是否直接包含图像文件
        """
        extract_path = Path(extract_dir)
        for ext in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.pdf']:
            if list(extract_path.glob(f'*{ext}')) or list(extract_path.glob(f'*{ext.upper()}')):
                return True
        return False
    
    def process_single_folder(self, folder_path: Path, folder_name: str) -> Dict:
        """
        处理单个文件夹的车辆档案
        
        Args:
            folder_path: 文件夹路径
            folder_name: 文件夹名称（用于显示）
            
        Returns:
            处理结果字典
        """
        print(f"\n处理文件夹: {folder_name}")
        print("-" * 70)
        
        # 处理文档
        documents = self.processor.process_folder(str(folder_path))
        
        if not documents:
            print("  警告: 未找到有效文档")
            return {
                'archive_name': folder_name,
                'documents': [],
                'merged_info': VehicleInfo(),
                'success': False
            }
        
        # 合并信息
        all_info = [doc.vehicle_info for doc in documents if doc.vehicle_info]
        merged_info = self.extractor.merge_info(*all_info) if all_info else VehicleInfo()
        
        print("\n  汇总信息:")
        if merged_info.vin:
            print(f"    车架号: {merged_info.vin}")
        if merged_info.owner_name:
            print(f"    车主: {merged_info.owner_name}")
        if merged_info.invoice_amount:
            print(f"    发票金额: {merged_info.invoice_amount} 元")
        if merged_info.plate_number:
            print(f"    车牌号: {merged_info.plate_number}")
        if merged_info.vehicle_model:
            print(f"    车辆型号: {merged_info.vehicle_model}")
        
        return {
            'archive_name': folder_name,
            'documents': documents,
            'merged_info': merged_info,
            'success': True
        }
    
    def process_archive(self, archive_path: str) -> List[Dict]:
        """
        处理单个压缩包（支持多文件夹结构）
        
        Args:
            archive_path: 压缩包路径
            
        Returns:
            处理结果列表（每个文件夹一个结果）
        """
        archive_path = Path(archive_path)
        print(f"\n处理压缩包: {archive_path.name}")
        print("=" * 70)
        
        # 解压
        import tempfile
        extract_dir = self.processor.extract_archive(str(archive_path))
        
        try:
            # 跳过中间文件夹层级（如果存在）
            working_dir = self._skip_intermediate_folder(extract_dir)
            
            # 检测是否有子文件夹
            subfolders = self._detect_subfolders(working_dir)
            has_root_images = self._has_images_in_root(working_dir)
            
            results = []
            
            if subfolders and not has_root_images:
                # 多文件夹模式
                print(f"检测到 {len(subfolders)} 个子文件夹，每个文件夹独立处理\n")
                
                for i, subfolder in enumerate(subfolders, 1):
                    print(f"[{i}/{len(subfolders)}] ", end='')
                    result = self.process_single_folder(subfolder, subfolder.name)
                    results.append(result)
            
            else:
                # 单一模式：所有文件在根目录
                print("单一档案模式（所有文件在根目录）\n")
                documents = self.processor.process_archive(str(archive_path))
                
                if not documents:
                    print("警告: 未找到有效文档")
                    return [{
                        'archive_name': archive_path.name,
                        'documents': [],
                        'merged_info': VehicleInfo(),
                        'success': False
                    }]
                
                # 合并信息
                all_info = [doc.vehicle_info for doc in documents if doc.vehicle_info]
                merged_info = self.extractor.merge_info(*all_info) if all_info else VehicleInfo()
                
                print("\n汇总信息:")
                print("-" * 70)
                if merged_info.vin:
                    print(f"  车架号: {merged_info.vin}")
                if merged_info.owner_name:
                    print(f"  车主: {merged_info.owner_name}")
                if merged_info.invoice_amount:
                    print(f"  发票金额: {merged_info.invoice_amount} 元")
                if merged_info.plate_number:
                    print(f"  车牌号: {merged_info.plate_number}")
                if merged_info.vehicle_model:
                    print(f"  车辆型号: {merged_info.vehicle_model}")
                
                results = [{
                    'archive_name': archive_path.name,
                    'documents': documents,
                    'merged_info': merged_info,
                    'success': True
                }]
            
            return results
            
        finally:
            # 清理临时目录
            if extract_dir.startswith(tempfile.gettempdir()):
                shutil.rmtree(extract_dir, ignore_errors=True)
    
    def organize_files(self, result: Dict, source_archive: str):
        """
        整理文件：创建以车架号命名的文件夹，重命名文件
        
        Args:
            result: 处理结果
            source_archive: 源压缩包路径
        """
        merged_info = result['merged_info']
        documents = result['documents']
        
        # 确定文件夹名称（优先使用车架号）
        if merged_info.vin:
            folder_name = merged_info.vin
        else:
            # 使用压缩包名称
            folder_name = Path(source_archive).stem
        
        # 创建目标文件夹
        target_folder = self.output_dir / folder_name
        target_folder.mkdir(exist_ok=True)
        
        print(f"\n整理文件到: {target_folder}")
        print("-" * 70)
        
        # 文档类型到中文名称的映射
        type_to_name = {
            'registration': '登记证',
            'invoice': '发票',
            'license': '行驶证',
        }
        
        # 复制并重命名文件
        copied_files = []
        for doc in documents:
            doc_type = doc.doc_type
            
            if doc_type in type_to_name:
                new_name = type_to_name[doc_type]
            else:
                # 未知类型，保留原文件名
                new_name = Path(doc.file_path).stem
            
            # 保留原扩展名
            ext = Path(doc.file_path).suffix
            new_filename = f"{new_name}{ext}"
            
            # 如果已存在同名文件，添加序号
            counter = 1
            final_path = target_folder / new_filename
            while final_path.exists():
                new_filename = f"{new_name}_{counter}{ext}"
                final_path = target_folder / new_filename
                counter += 1
            
            # 复制文件
            try:
                shutil.copy2(doc.file_path, final_path)
                print(f"  ✓ {Path(doc.file_path).name} → {new_filename}")
                copied_files.append(str(final_path))
            except Exception as e:
                print(f"  ✗ 复制失败: {e}")
        
        return target_folder, copied_files
    
    def export_to_excel(self, results: List[Dict], excel_path: str = None):
        """
        导出结果到Excel（每个档案一行，区分不同来源的字段）
        
        Args:
            results: 处理结果列表
            excel_path: Excel文件路径
        """
        if excel_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = self.output_dir / f"车辆信息汇总_{timestamp}.xlsx"
        else:
            excel_path = Path(excel_path)
        
        print(f"\n导出Excel: {excel_path.name}")
        print("-" * 70)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "车辆信息"
        
        # 设置表头 - 简化字段
        headers = [
            '序号', '来源文件',
            '车架号(行驶证)', '车架号(登记证)', '车架号(发票)',
            '车主(行驶证)', '新车主(登记证)', '交易金额(发票)'
        ]
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据 - 每个档案一行
        row_num = 1
        for result in results:
            if not result['success']:
                continue
            
            archive_name = result['archive_name']
            documents = result['documents']
            
            # 分类存储不同来源的信息
            license_info = None  # 行驶证
            registration_info = None  # 登记证（注册登记页）
            registration_transfer_info = None  # 登记证尾页（转移登记）
            invoice_info = None  # 发票
            
            for doc in documents:
                if doc.doc_type == 'license':
                    license_info = doc.vehicle_info
                elif doc.doc_type == 'registration':
                    registration_info = doc.vehicle_info
                elif doc.doc_type == 'registration_transfer':
                    registration_transfer_info = doc.vehicle_info
                elif doc.doc_type == 'invoice':
                    invoice_info = doc.vehicle_info
            
            # 构建一行数据
            row_num += 1
            
            # 车架号(登记证)：只有在识别出来时才显示
            vin_registration = registration_info.vin if registration_info and registration_info.vin else ''
            
            # 新车主：优先从登记证尾页提取，其次从登记证
            new_owner = ''
            if registration_transfer_info and registration_transfer_info.new_owner_name:
                new_owner = registration_transfer_info.new_owner_name
            elif registration_info and registration_info.new_owner_name:
                new_owner = registration_info.new_owner_name
            
            row = [
                row_num - 1,  # 序号
                archive_name,  # 来源文件
                license_info.vin if license_info else '',  # 车架号(行驶证)
                vin_registration,  # 车架号(登记证)，未识别时为空
                invoice_info.vin if invoice_info else '',  # 车架号(发票)
                license_info.owner_name if license_info else '',  # 车主(行驶证)
                new_owner,  # 新车主(登记证/登记证尾页)
                invoice_info.invoice_amount if invoice_info else '',  # 交易金额(发票)
            ]
            
            ws.append(row)
            
            # vin_registration 已在上面定义
            vin_license = license_info.vin if license_info else ''
            vin_registration = registration_info.vin if registration_info else ''
            
            # 如果两个车架号都存在且不一致，标红（注意：row_num 在 append 之前已经 +1 了）
            if vin_license and vin_registration and vin_license != vin_registration:
                red_font = Font(color="FF0000", bold=True)  # 红色粗体
                red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # 浅红色背景
                
                # 标红车架号(行驶证)列 - 第3列(C)
                cell_license = ws.cell(row_num, 3)
                cell_license.font = red_font
                cell_license.fill = red_fill
                
                # 标红车架号(登记证)列 - 第4列(D)
                cell_registration = ws.cell(row_num, 4)
                cell_registration.font = red_font
                cell_registration.fill = red_fill
                
                print(f"  {row_num - 1}. {archive_name} | ⚠️  车架号不一致！")
                print(f"      行驶证: {vin_license}")
                print(f"      登记证: {vin_registration}")
            else:
                # 优先显示行驶证或登记证的VIN，否则显示发票的VIN
                vin_display = vin_license or vin_registration or (invoice_info.vin if invoice_info else '') or '未识别'
                print(f"  {row_num - 1}. {archive_name} | VIN: {vin_display}")
        
        # 调整列宽（增加车主和新车主列宽，适应长公司名）
        column_widths = [6, 20, 22, 22, 22, 30, 30, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width
        
        # 设置数据行的对齐方式和自动换行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='left', 
                    vertical='center',
                    wrap_text=True  # 启用自动换行
                )
        
        # 保存
        wb.save(excel_path)
        print(f"\n✓ Excel导出成功: {excel_path}")
        print(f"✓ 车架号不一致的记录已用红色标记")
        
        return excel_path
    
    def process_multiple_archives(self, archive_paths: List[str], 
                                  organize: bool = True,
                                  export_excel: bool = True):
        """
        批量处理多个压缩包
        
        Args:
            archive_paths: 压缩包路径列表
            organize: 是否整理文件
            export_excel: 是否导出Excel
        """
        print("\n" + "=" * 70)
        print("车辆文档批处理")
        print("=" * 70)
        print(f"压缩包数量: {len(archive_paths)}")
        print(f"输出目录: {self.output_dir.absolute()}")
        print("=" * 70)
        
        results = []
        
        for i, archive_path in enumerate(archive_paths, 1):
            print(f"\n[压缩包 {i}/{len(archive_paths)}]")
            
            try:
                # 处理压缩包（现在返回列表）
                archive_results = self.process_archive(archive_path)
                
                # 每个文件夹的结果
                for result in archive_results:
                    results.append(result)
                    
                    # 整理文件
                    if organize and result['success']:
                        self.organize_files(result, archive_path)
                
            except Exception as e:
                print(f"✗ 处理失败: {e}")
                import traceback
                traceback.print_exc()
        
        # 导出Excel
        if export_excel and results:
            try:
                self.export_to_excel(results)
            except Exception as e:
                print(f"\n✗ Excel导出失败: {e}")
        
        # 打印总结
        print("\n" + "=" * 70)
        print("处理完成")
        print("=" * 70)
        
        success_count = sum(1 for r in results if r['success'])
        print(f"成功: {success_count}/{len(results)}")
        print(f"输出目录: {self.output_dir.absolute()}")
        
        return results


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="车辆文档批处理工具 - 自动识别、整理、导出Excel"
    )
    parser.add_argument('archives', nargs='+', help='压缩包文件路径（支持多个）')
    parser.add_argument('-o', '--output', default='output',
                       help='输出目录 (默认: output)')
    parser.add_argument('-e', '--engine', default='paddleocr',
                       choices=['paddleocr', 'rapidocr', 'easyocr'],
                       help='OCR引擎 (默认: paddleocr)')
    parser.add_argument('--no-organize', action='store_true',
                       help='不整理文件')
    parser.add_argument('--no-excel', action='store_true',
                       help='不导出Excel')
    parser.add_argument('-v', '--verbose', action='store_true',
                       help='显示详细日志（DEBUG级别）')
    parser.add_argument('--debug', action='store_true',
                       help='显示调试日志（包括每个正则匹配尝试）')
    
    args = parser.parse_args()
    
    # 配置日志
    if args.debug:
        log_level = logging.DEBUG
        log_format = '%(levelname)s - %(name)s - %(message)s'
    elif args.verbose:
        log_level = logging.INFO
        log_format = '%(levelname)s - %(message)s'
    else:
        log_level = logging.WARNING
        log_format = '%(message)s'
    
    logging.basicConfig(
        level=log_level,
        format=log_format,
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.info("车辆文档批处理工具启动")
    
    # 检查文件是否存在
    for archive in args.archives:
        if not Path(archive).exists():
            print(f"错误: 文件不存在: {archive}")
            sys.exit(1)
    
    # 创建处理器
    processor = VehicleArchiveProcessor(
        output_dir=args.output,
        ocr_engine=args.engine
    )
    
    # 批量处理
    processor.process_multiple_archives(
        args.archives,
        organize=not args.no_organize,
        export_excel=not args.no_excel
    )


if __name__ == '__main__':
    main()
