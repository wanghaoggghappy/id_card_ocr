#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试Excel导出行号修复
"""

from pathlib import Path
from openpyxl import load_workbook
from vehicle_info_extractor import VehicleInfo
from process_vehicle_archives import VehicleArchiveProcessor

# 模拟测试数据
def create_test_results():
    """创建测试结果数据"""
    results = []
    
    # 档案1: VIN一致，无标红
    license_info1 = VehicleInfo(
        vin='LSVNP60C0PN046761',
        owner_name='刘顺',
        plate_number='云A1M57C'
    )
    registration_info1 = VehicleInfo(
        vin='LSVNP60C0PN046761'
    )
    invoice_info1 = VehicleInfo(
        vin='LSVNP60C0PN046761',
        invoice_amount='45000.00'
    )
    
    # 档案2: VIN一致，无标红
    license_info2 = VehicleInfo(
        vin='LSVNP60C8PN049942',
        owner_name='成都八保舟汽车服务有限公司',
        plate_number='云A9VQ21'
    )
    invoice_info2 = VehicleInfo(
        vin='LSVXP60C8PN049942',
        invoice_amount='45000.00'
    )
    
    # 档案3: VIN一致，无标红
    license_info3 = VehicleInfo(
        vin='LSVNP60CXPN048209',
        owner_name='成都八保舟汽车服务有限公司',
        plate_number='云A2M58D'
    )
    invoice_info3 = VehicleInfo(
        vin='LSVNP60CXPN048209',
        invoice_amount='45000.00'
    )
    
    # 档案4: VIN不一致，应标红第4行
    license_info4 = VehicleInfo(
        vin='LSVNP60C1PN048194',
        owner_name='成都八保舟汽车服务有限公司',
        plate_number='云A3M59E'
    )
    registration_info4 = VehicleInfo(
        vin='LSVNP60C1PN048194'
    )
    invoice_info4 = VehicleInfo(
        vin='LSVNP60C1PN048194',
        invoice_amount='45000.00'
    )
    
    # 档案5: VIN不一致，应标红第5行（这是用户截图中的红框位置，但应该是第4行）
    license_info5 = VehicleInfo(
        vin='LSVNF60C8NN097361',
        owner_name='昆明福超汽车销售有限公司',
        plate_number='云A4M60F'
    )
    registration_info5 = VehicleInfo(
        vin='LSVNF68C8NN097361'  # 故意不一致
    )
    invoice_info5 = VehicleInfo(
        vin='LSVNF68C8NN097361',
        invoice_amount='45000.00'
    )
    
    # 构建结果结构（模拟 process_archive 返回的格式）
    from dataclasses import dataclass
    from typing import Optional
    
    @dataclass
    class MockDocument:
        doc_type: str
        vehicle_info: VehicleInfo
    
    results.append({
        'success': True,
        'archive_name': 'LSVNP60C0PN046761',
        'documents': [
            MockDocument('license', license_info1),
            MockDocument('registration', registration_info1),
            MockDocument('invoice', invoice_info1)
        ]
    })
    
    results.append({
        'success': True,
        'archive_name': 'LSVNP60C8PN049942',
        'documents': [
            MockDocument('license', license_info2),
            MockDocument('invoice', invoice_info2)
        ]
    })
    
    results.append({
        'success': True,
        'archive_name': 'LSVNP60CXPN048209',
        'documents': [
            MockDocument('license', license_info3),
            MockDocument('invoice', invoice_info3)
        ]
    })
    
    results.append({
        'success': True,
        'archive_name': 'LSVNP60C1PN048194',
        'documents': [
            MockDocument('license', license_info4),
            MockDocument('registration', registration_info4),
            MockDocument('invoice', invoice_info4)
        ]
    })
    
    results.append({
        'success': True,
        'archive_name': 'LSVNF60C8NN097361',
        'documents': [
            MockDocument('license', license_info5),
            MockDocument('registration', registration_info5),
            MockDocument('invoice', invoice_info5)
        ]
    })
    
    return results

def test_excel_export():
    """测试Excel导出和行号修复"""
    print("="*70)
    print("测试Excel导出行号修复")
    print("="*70)
    
    # 创建处理器
    processor = VehicleArchiveProcessor()
    
    # 生成测试数据
    results = create_test_results()
    
    # 导出Excel
    excel_path = processor.output_dir / "test_excel_row_fix.xlsx"
    processor.export_to_excel(results, excel_path)
    
    # 验证结果
    print("\n验证Excel内容:")
    print("-"*70)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 检查标红的行
    red_rows = []
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row_num, 3)  # 车架号(行驶证)列
        if cell.font and cell.font.color and str(cell.font.color.rgb) == 'FFFF0000':
            red_rows.append(row_num)
    
    print(f"\n标红的行号: {red_rows}")
    
    # 预期：第5行应该标红（数据行从第2行开始，第5行是第4个数据档案）
    if 5 in red_rows:
        print("✓ 第5行正确标红（VIN不一致）")
    else:
        print(f"✗ 第5行未标红（实际标红: {red_rows}）")
    
    # 检查第5行数据
    print("\n第5行数据:")
    for col in range(1, 9):
        value = ws.cell(5, col).value
        print(f"  列{col}: {value}")
    
    print(f"\n✓ 测试完成，Excel保存至: {excel_path}")
    
if __name__ == "__main__":
    test_excel_export()
